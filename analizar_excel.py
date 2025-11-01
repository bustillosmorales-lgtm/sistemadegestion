"""
Script para analizar la estructura del archivo Gestión Full3.xlsm
"""

import openpyxl
import sys

# Abrir el archivo
archivo = r"C:\Users\franc\OneDrive-mail.udp.cl\Documentos\sistema\nuevo_sistema\Gestión Full3.xlsm"

try:
    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)

    print("="*80)
    print("ANÁLISIS DE ESTRUCTURA: Gestión Full3.xlsm")
    print("="*80)
    print(f"\nTotal de hojas: {len(wb.sheetnames)}\n")

    for sheet_name in wb.sheetnames:
        print(f"\n{'='*80}")
        print(f"HOJA: {sheet_name}")
        print(f"{'='*80}")

        ws = wb[sheet_name]

        # Obtener dimensiones
        max_row = ws.max_row
        max_col = ws.max_column

        print(f"Dimensiones: {max_row} filas x {max_col} columnas")

        # Leer encabezados (primera fila)
        if max_row > 0:
            print("\nENCABEZADOS (Primera fila):")
            headers = []
            for col in range(1, min(max_col + 1, 30)):  # Máximo 30 columnas
                cell_value = ws.cell(1, col).value
                if cell_value:
                    headers.append(f"  Col {col}: {cell_value}")

            for header in headers[:20]:  # Mostrar máximo 20 encabezados
                print(header)

            if len(headers) > 20:
                print(f"  ... y {len(headers) - 20} columnas más")

            # Mostrar muestra de datos (filas 2-6)
            print("\nMUESTRA DE DATOS (filas 2-6):")
            for row_idx in range(2, min(7, max_row + 1)):
                print(f"\n  Fila {row_idx}:")
                for col_idx in range(1, min(max_col + 1, 10)):  # Máximo 10 columnas
                    cell_value = ws.cell(row_idx, col_idx).value
                    header = ws.cell(1, col_idx).value
                    if cell_value is not None:
                        print(f"    {header}: {cell_value}")
        else:
            print("  (Hoja vacía)")

    wb.close()

    print("\n" + "="*80)
    print("ANÁLISIS COMPLETADO")
    print("="*80)

except Exception as e:
    print(f"ERROR: {e}")
    import traceback
    traceback.print_exc()
