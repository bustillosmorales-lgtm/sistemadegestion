"""
Script para cargar datos desde Excel a Supabase
Sin límites de tiempo - corre localmente en tu PC
"""

import openpyxl
import requests
import os
from datetime import datetime
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv('.env.local')

SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_SERVICE_KEY = os.getenv('SUPABASE_SERVICE_KEY')

if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
    print("❌ Error: Faltan variables SUPABASE_URL o SUPABASE_SERVICE_KEY en .env.local")
    exit(1)

HEADERS = {
    'apikey': SUPABASE_SERVICE_KEY,
    'Authorization': f'Bearer {SUPABASE_SERVICE_KEY}',
    'Content-Type': 'application/json',
    'Prefer': 'return=minimal'
}

def excel_date_to_iso(excel_date):
    """Convierte fecha de Excel a formato ISO"""
    if excel_date is None:
        return None

    if isinstance(excel_date, str):
        try:
            date = datetime.strptime(excel_date, '%Y-%m-%d')
            return date.strftime('%Y-%m-%d')
        except:
            return None

    if isinstance(excel_date, datetime):
        return excel_date.strftime('%Y-%m-%d')

    # Número de Excel (días desde 1900-01-01)
    if isinstance(excel_date, (int, float)):
        base_date = datetime(1899, 12, 30)
        delta_days = int(excel_date)
        date = base_date + timedelta(days=delta_days)
        return date.strftime('%Y-%m-%d')

    return None

def insertar_batch(tabla, datos, batch_size=50):
    """Inserta datos en lotes a Supabase"""
    if not datos:
        return 0

    total = len(datos)
    insertados = 0

    for i in range(0, total, batch_size):
        batch = datos[i:i+batch_size]

        url = f"{SUPABASE_URL}/rest/v1/{tabla}"
        response = requests.post(url, json=batch, headers=HEADERS)

        if response.status_code in [200, 201]:
            insertados += len(batch)
            print(f"  ✓ {insertados}/{total} registros insertados...")
        else:
            print(f"  ❌ Error insertando batch: {response.status_code}")
            print(f"     {response.text}")
            return insertados

    return insertados

def procesar_ventas(archivo_excel):
    """Procesa la hoja de ventas"""
    print("\n📊 Procesando VENTAS...")

    wb = openpyxl.load_workbook(archivo_excel, data_only=True)

    if 'ventas' not in wb.sheetnames:
        print("  ⚠️  Hoja 'ventas' no encontrada")
        return

    ws = wb['ventas']
    registros = []

    # Saltar header (fila 1)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row_idx % 1000 == 0:
            print(f"  Leyendo fila {row_idx}...")

        empresa = str(row[0]).strip() if row[0] else None
        canal = str(row[1]).strip() if row[1] else None

        # Solo TLT + MELI
        if not empresa or not canal:
            continue
        if empresa.upper() != 'TLT' or canal.upper() != 'MELI':
            continue

        sku = str(row[19]).strip() if row[19] else None
        unidades = float(row[10]) if row[10] else 0

        if not sku or unidades <= 0:
            continue

        registros.append({
            'empresa': empresa,
            'canal': canal,
            'fecha': excel_date_to_iso(row[5]),
            'unidades': unidades,
            'sku': sku,
            'mlc': str(row[20]).strip() if row[20] else '',
            'descripcion': str(row[21]).strip() if row[21] else '',
            'precio': float(row[23]) if row[23] else 0
        })

    print(f"  📝 {len(registros)} ventas válidas encontradas")

    if registros:
        insertados = insertar_batch('ventas_historicas', registros)
        print(f"  ✅ {insertados} ventas insertadas en Supabase")

    wb.close()

def procesar_stock(archivo_excel):
    """Procesa la hoja de Stock"""
    print("\n📦 Procesando STOCK...")

    wb = openpyxl.load_workbook(archivo_excel, data_only=True)

    if 'Stock' not in wb.sheetnames:
        print("  ⚠️  Hoja 'Stock' no encontrada")
        return

    ws = wb['Stock']
    registros = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[0]).strip() if row[0] else None

        if not sku:
            continue

        registros.append({
            'sku': sku,
            'descripcion': str(row[1]).strip() if row[1] else '',
            'bodega_c': float(row[2]) if row[2] else 0,
            'bodega_d': float(row[3]) if row[3] else 0,
            'bodega_e': float(row[4]) if row[4] else 0,
            'bodega_f': float(row[5]) if row[5] else 0,
            'bodega_h': float(row[7]) if row[7] else 0,
            'bodega_j': float(row[9]) if row[9] else 0
        })

    print(f"  📝 {len(registros)} SKUs encontrados")

    if registros:
        # Usar upsert para stock
        url = f"{SUPABASE_URL}/rest/v1/stock_actual"
        headers_upsert = {**HEADERS, 'Prefer': 'resolution=merge-duplicates'}

        for i in range(0, len(registros), 50):
            batch = registros[i:i+50]
            response = requests.post(url, json=batch, headers=headers_upsert)
            if response.status_code in [200, 201]:
                print(f"  ✓ {min(i+50, len(registros))}/{len(registros)} SKUs insertados...")

        print(f"  ✅ {len(registros)} SKUs insertados en Supabase")

    wb.close()

def procesar_transito(archivo_excel):
    """Procesa la hoja de tránsito china"""
    print("\n🚢 Procesando TRÁNSITO CHINA...")

    wb = openpyxl.load_workbook(archivo_excel, data_only=True)

    if 'transito china' not in wb.sheetnames:
        print("  ⚠️  Hoja 'transito china' no encontrada")
        return

    ws = wb['transito china']
    registros = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[3]).strip() if row[3] else None
        unidades = float(row[7]) if row[7] else 0

        if not sku or unidades <= 0:
            continue

        registros.append({
            'sku': sku,
            'unidades': unidades,
            'estado': 'en_transito'
        })

    print(f"  📝 {len(registros)} registros encontrados")

    if registros:
        insertados = insertar_batch('transito_china', registros)
        print(f"  ✅ {insertados} registros insertados en Supabase")

    wb.close()

def procesar_compras(archivo_excel):
    """Procesa la hoja de compras"""
    print("\n🛒 Procesando COMPRAS...")

    wb = openpyxl.load_workbook(archivo_excel, data_only=True)

    if 'compras' not in wb.sheetnames:
        print("  ⚠️  Hoja 'compras' no encontrada")
        return

    ws = wb['compras']
    registros = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[0]).strip() if row[0] else None
        fecha = excel_date_to_iso(row[3])

        if not sku or not fecha:
            continue

        registros.append({
            'sku': sku,
            'fecha_compra': fecha
        })

    print(f"  📝 {len(registros)} compras encontradas")

    if registros:
        insertados = insertar_batch('compras_historicas', registros)
        print(f"  ✅ {insertados} compras insertadas en Supabase")

    wb.close()

def procesar_packs(archivo_excel):
    """Procesa la hoja de Packs"""
    print("\n📦 Procesando PACKS...")

    wb = openpyxl.load_workbook(archivo_excel, data_only=True)

    if 'Packs' not in wb.sheetnames:
        print("  ⚠️  Hoja 'Packs' no encontrada")
        return

    ws = wb['Packs']
    registros = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        sku_pack = str(row[0]).strip() if row[0] else None
        sku_componente = str(row[1]).strip() if row[1] else None
        cantidad = float(row[2]) if row[2] else 1

        if not sku_pack or not sku_componente:
            continue

        registros.append({
            'sku_pack': sku_pack,
            'sku_componente': sku_componente,
            'cantidad': cantidad
        })

    print(f"  📝 {len(registros)} packs encontrados")

    if registros:
        insertados = insertar_batch('packs', registros)
        print(f"  ✅ {insertados} packs insertados en Supabase")

    wb.close()

def main():
    print("="*60)
    print("📊 CARGA DE DATOS EXCEL A SUPABASE")
    print("="*60)

    # Verificar que existan los archivos
    archivo_ventas = 'ventas.xlsx'
    archivo_otros = 'otros_datos.xlsx'

    if not os.path.exists(archivo_ventas):
        print(f"\n❌ No se encontró el archivo '{archivo_ventas}'")
        print("   Crea este archivo con solo la hoja 'ventas'")
        return

    if not os.path.exists(archivo_otros):
        print(f"\n❌ No se encontró el archivo '{archivo_otros}'")
        print("   Crea este archivo con Stock, transito china, compras, Packs")
        return

    print(f"\n✅ Archivos encontrados:")
    print(f"   - {archivo_ventas}")
    print(f"   - {archivo_otros}")

    print(f"\n🔗 Conectando a Supabase: {SUPABASE_URL}")

    # Procesar archivos
    procesar_ventas(archivo_ventas)
    procesar_stock(archivo_otros)
    procesar_transito(archivo_otros)
    procesar_compras(archivo_otros)
    procesar_packs(archivo_otros)

    print("\n" + "="*60)
    print("🎉 ¡PROCESO COMPLETADO!")
    print("="*60)
    print("\nVerifica los datos en Supabase:")
    print(f"{SUPABASE_URL.replace('/rest/v1', '')}/project/default/editor")

if __name__ == '__main__':
    from datetime import timedelta
    main()
