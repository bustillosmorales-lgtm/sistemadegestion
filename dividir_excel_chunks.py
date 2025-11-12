"""
Script para dividir Excel grande en chunks pequeños
Cada chunk puede ser procesado por Netlify Functions (<10 segundos)
"""

import openpyxl
import os
from datetime import datetime

# Configuración
FILAS_POR_CHUNK = 2000  # Aprox 2000 filas = ~5 segundos de procesamiento
OUTPUT_DIR = 'chunks_excel'

def dividir_hoja_ventas(archivo_entrada):
    """Divide la hoja de ventas en múltiples archivos"""
    print(f"\n📊 Dividiendo hoja VENTAS de {archivo_entrada}...")

    wb = openpyxl.load_workbook(archivo_entrada, data_only=True)

    if 'ventas' not in wb.sheetnames:
        print("  ⚠️  Hoja 'ventas' no encontrada")
        wb.close()
        return []

    ws_original = wb['ventas']

    # Obtener header (fila 1)
    header = [cell.value for cell in ws_original[1]]

    # Contar filas totales
    total_filas = ws_original.max_row - 1  # -1 porque excluimos header
    print(f"  📝 Total filas: {total_filas}")

    # Calcular chunks necesarios
    num_chunks = (total_filas // FILAS_POR_CHUNK) + (1 if total_filas % FILAS_POR_CHUNK else 0)
    print(f"  📦 Se crearán {num_chunks} chunks de ventas")

    archivos_generados = []
    fila_inicio = 2  # Empezar después del header

    for chunk_num in range(1, num_chunks + 1):
        # Crear nuevo workbook para este chunk
        wb_chunk = openpyxl.Workbook()
        ws_chunk = wb_chunk.active
        ws_chunk.title = 'ventas'

        # Copiar header
        ws_chunk.append(header)

        # Copiar filas para este chunk
        fila_fin = min(fila_inicio + FILAS_POR_CHUNK - 1, ws_original.max_row)
        filas_copiadas = 0

        for row_idx in range(fila_inicio, fila_fin + 1):
            row_values = [cell.value for cell in ws_original[row_idx]]
            ws_chunk.append(row_values)
            filas_copiadas += 1

        # Guardar chunk
        nombre_archivo = f'ventas_chunk_{chunk_num:02d}.xlsx'
        ruta_completa = os.path.join(OUTPUT_DIR, nombre_archivo)
        wb_chunk.save(ruta_completa)
        archivos_generados.append(nombre_archivo)

        print(f"  ✓ Chunk {chunk_num}/{num_chunks}: {filas_copiadas} filas → {nombre_archivo}")

        wb_chunk.close()
        fila_inicio = fila_fin + 1

    wb.close()
    return archivos_generados

def copiar_otras_hojas(archivo_entrada):
    """Copia las hojas pequeñas (Stock, transito, compras, packs) a un solo archivo"""
    print(f"\n📦 Copiando hojas auxiliares de {archivo_entrada}...")

    wb = openpyxl.load_workbook(archivo_entrada, data_only=True)

    # Hojas a copiar (todo excepto ventas)
    hojas_copiar = ['Stock', 'transito china', 'compras', 'Packs', 'desconsiderar']
    hojas_encontradas = [h for h in hojas_copiar if h in wb.sheetnames]

    if not hojas_encontradas:
        print("  ⚠️  No se encontraron hojas auxiliares")
        wb.close()
        return None

    # Crear nuevo workbook
    wb_nuevo = openpyxl.Workbook()
    wb_nuevo.remove(wb_nuevo.active)  # Remover hoja default

    for nombre_hoja in hojas_encontradas:
        ws_original = wb[nombre_hoja]
        ws_nueva = wb_nuevo.create_sheet(nombre_hoja)

        # Copiar todas las filas
        filas_copiadas = 0
        for row in ws_original.iter_rows(values_only=True):
            ws_nueva.append(row)
            filas_copiadas += 1

        print(f"  ✓ Hoja '{nombre_hoja}': {filas_copiadas} filas copiadas")

    # Guardar
    nombre_archivo = 'otros_datos.xlsx'
    ruta_completa = os.path.join(OUTPUT_DIR, nombre_archivo)
    wb_nuevo.save(ruta_completa)

    wb.close()
    wb_nuevo.close()

    print(f"  ✅ Hojas auxiliares guardadas en: {nombre_archivo}")
    return nombre_archivo

def generar_instrucciones(archivos_ventas, archivo_otros):
    """Genera un archivo con instrucciones de carga"""
    instrucciones = """
╔════════════════════════════════════════════════════════════════╗
║          INSTRUCCIONES PARA CARGAR DATOS A SUPABASE            ║
╔════════════════════════════════════════════════════════════════╗

📋 ARCHIVOS GENERADOS:
"""

    if archivo_otros:
        instrucciones += f"\n1. {archivo_otros} (Stock, tránsito, compras, packs)\n"

    for i, archivo in enumerate(archivos_ventas, start=2):
        instrucciones += f"{i}. {archivo}\n"

    instrucciones += """
📤 ORDEN DE CARGA:

IMPORTANTE: Sube los archivos EN ESTE ORDEN en https://sistemadegestion.net

"""

    if archivo_otros:
        instrucciones += f"PRIMERO: {archivo_otros}\n"

    for i, archivo in enumerate(archivos_ventas, start=1):
        instrucciones += f"Chunk {i}: {archivo}\n"

    instrucciones += """

⏱️ TIEMPO ESTIMADO:
- Cada archivo tarda ~5-8 segundos en procesarse
- Total: """ + f"{(len(archivos_ventas) + 1) * 7} segundos aprox.\n"

    instrucciones += """

✅ VERIFICACIÓN:
Después de cargar todos los archivos, verifica en Supabase:
https://supabase.com/dashboard/project/ugabltnuwwtbpyqoptdg/editor

Revisa las tablas:
- ventas_historicas (debería tener miles de registros)
- stock_actual
- transito_china
- compras_historicas
- packs

🔄 PARA CARGAS DIARIAS:
En el futuro, solo necesitarás cargar:
1. Las ventas del día anterior (un archivo pequeño)
2. El stock actualizado
"""

    # Guardar instrucciones
    ruta_instrucciones = os.path.join(OUTPUT_DIR, 'INSTRUCCIONES_CARGA.txt')
    with open(ruta_instrucciones, 'w', encoding='utf-8') as f:
        f.write(instrucciones)

    print("\n" + instrucciones)

def main():
    print("="*70)
    print("  📊 DIVIDIR EXCEL EN CHUNKS PARA NETLIFY FUNCTIONS")
    print("="*70)

    # Crear directorio de salida
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        print(f"\n✅ Directorio creado: {OUTPUT_DIR}/")

    # Buscar archivo de entrada
    archivo_entrada = None
    for nombre in ['tablas_base.xlsx', 'ventas.xlsx', 'datos.xlsx']:
        if os.path.exists(nombre):
            archivo_entrada = nombre
            break

    if not archivo_entrada:
        print("\n❌ ERROR: No se encontró archivo Excel")
        print("\nColoca tu archivo Excel en esta carpeta con alguno de estos nombres:")
        print("  - tablas_base.xlsx")
        print("  - ventas.xlsx")
        print("  - datos.xlsx")
        return

    print(f"\n✅ Archivo encontrado: {archivo_entrada}")
    print(f"   Tamaño: {os.path.getsize(archivo_entrada) / 1024 / 1024:.2f} MB")

    # Dividir ventas
    archivos_ventas = dividir_hoja_ventas(archivo_entrada)

    # Copiar otras hojas
    archivo_otros = copiar_otras_hojas(archivo_entrada)

    # Generar instrucciones
    generar_instrucciones(archivos_ventas, archivo_otros)

    print("\n" + "="*70)
    print("  🎉 ¡PROCESO COMPLETADO!")
    print("="*70)
    print(f"\n📁 Archivos generados en: {os.path.abspath(OUTPUT_DIR)}/")
    print(f"\n📋 Sigue las instrucciones en: {OUTPUT_DIR}/INSTRUCCIONES_CARGA.txt")
    print("\n💡 Abre https://sistemadegestion.net y sube los archivos en orden")

if __name__ == '__main__':
    main()
